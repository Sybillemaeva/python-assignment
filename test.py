from openpyxl import load_workbook
import pandas as pd
import numpy

filename = "C:\\Users\DELL\OneDrive\Bureau\Python Assignment\PYTHON ASSIGMENT.xlsx"

wb = load_workbook(filename)

sheet = wb['PYTHON ASSIGMENT']

emails = []

modified_emails = []

def extract_emails():
    for x in range(sheet._current_row):
        if(x == 0):
            #print('Jumping the headers')
            print(' ')
        else:
            mail = sheet.cell(row=x + 1, column=2).value
            emails.append(mail)

def modify_emails():
    length = len(emails)
    for x in range(length):
        mail = emails[x].split("@")
        result = mail[0] + '@handsinhands.org'
        modified_emails.append(result)

def modify_xlsx():
    for x in range(sheet._current_row):
        if(x == 0):
            #print('Jumping the headers')
            print(' ')
        else:
            sheet.cell(row=x + 1, column=2).value = modified_emails[x-1]
            wb.save(filename)

def modify_csv():
    dx = pd.read_csv(r'C:\\Users\DELL\OneDrive\Bureau\Python Assignment\PYTHON ASSIGMENT.csv')
    size = len(modified_emails)
    for x in range(size):
        dx.loc[x, 'email'] = modified_emails[x]
        dx.to_csv('C:\\Users\DELL\OneDrive\Bureau\Python Assignment\PYTHON ASSIGMENT.csv', index=False)

def display_xlsx_file():
    df = pd.read_excel(r'C:\Users\DELL\OneDrive\Bureau\Python Assignment\PYTHON ASSIGMENT.xlsx')
    print(df)

def display_csv_file():
    df = pd.read_csv(r'C:\\Users\DELL\OneDrive\Bureau\Python Assignment\PYTHON ASSIGMENT.csv')
    print(df)


print("Extracting the emails to be modified")

extract_emails()

print("Modifying the emails...")

modify_emails()

print(' ')

print("Modifying both the .xlsx and .csv files and updating thier emails")

print(' ')

print("Before modification")

print(' ')

display_xlsx_file()

print(' ')

display_csv_file()

print(' ')

print('Modifying...')

modify_xlsx()

modify_csv()

print("After modification")

display_xlsx_file()

display_csv_file()
